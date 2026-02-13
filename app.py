"""
在庫Aging分析ツール — ユアトレード EC在庫向け
Streamlit Webアプリ (1ファイル構成)
"""

import io
import json
import os
import re
from datetime import datetime, timedelta
from urllib.error import URLError
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
AGING_BINS = [
    (0, 30, "0-30日"),
    (31, 60, "31-60日"),
    (61, 90, "61-90日"),
    (91, 180, "91-180日"),
    (181, 365, "181-365日"),
    (366, 999999, "365日超"),
]

SHOPEE_COLUMNS = [
    "Product ID", "Product Name", "Variation ID", "Variation Name",
    "Parent SKU", "SKU", "Price", "GTIN", "Stock",
    "Min Purchase Qty", "Fail Reason",
]

# Excel スタイル
FILL_SHOPEE = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
FILL_EXPIRED = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
FONT_EXPIRED = Font(color="FFFFFF", bold=True)
FILL_NEAR_EXPIRY = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_PINK = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# カスタム CSS
# ---------------------------------------------------------------------------
CUSTOM_CSS = """
<style>
/* ============================================================
   Light Tech Theme — 白/青/緑 クリーンなビジネステーマ
   ============================================================ */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Noto+Sans+JP:wght@300;400;500;700&display=swap');

/* --- ベース: ライト背景 --- */
.stApp {
    font-family: 'Inter', 'Noto Sans JP', sans-serif;
    background: #f5f7fa !important;
    color: #1e293b !important;
    font-size: 0.925rem;
}
.stApp > header { background: transparent !important; }
.stMarkdown, .stMarkdown p, .stMarkdown li,
.stCaption, label, .stSelectbox label, .stMultiSelect label {
    color: #475569 !important;
    font-size: 0.9rem !important;
}
h1, h2, h3, h4, h5, h6 { color: #0f172a !important; }

/* --- サイドバー --- */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #ffffff 0%, #f0f4f8 100%) !important;
    border-right: 1px solid #e2e8f0;
}
section[data-testid="stSidebar"] * {
    color: #334155 !important;
}
section[data-testid="stSidebar"] .stButton > button {
    background: #16a34a !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    padding: 0.6rem 1.2rem !important;
    letter-spacing: 0.02em !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 2px 8px rgba(22,163,74,0.25) !important;
}
section[data-testid="stSidebar"] .stButton > button * {
    color: #ffffff !important;
    background: transparent !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: #15803d !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(22,163,74,0.35) !important;
}

/* --- メインボタン --- */
.stApp .stButton > button[kind="primary"],
.stApp button[data-testid="stDownloadButton"] {
    background: #2563eb !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.2) !important;
    transition: all 0.2s ease !important;
}
.stApp .stButton > button[kind="primary"] *,
.stApp button[data-testid="stDownloadButton"] * {
    color: #ffffff !important;
    background: transparent !important;
}
.stApp .stButton > button[kind="primary"]:hover,
.stApp button[data-testid="stDownloadButton"]:hover {
    background: #1d4ed8 !important;
    box-shadow: 0 4px 16px rgba(37,99,235,0.3) !important;
    transform: translateY(-1px) !important;
}

/* --- KPI カード --- */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1.5rem 0 2rem 0;
}
.kpi-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 1.3rem 1.4rem;
    position: relative;
    overflow: hidden;
    transition: all 0.25s ease;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
}
.kpi-card:hover {
    transform: translateY(-3px);
    box-shadow: 0 8px 24px rgba(0,0,0,0.08);
}
.kpi-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.kpi-card.v1::before { background: linear-gradient(90deg, #2563eb, #3b82f6); }
.kpi-card.v2::before { background: linear-gradient(90deg, #0891b2, #06b6d4); }
.kpi-card.v3::before { background: linear-gradient(90deg, #dc2626, #ef4444); }
.kpi-card.v4::before { background: linear-gradient(90deg, #16a34a, #22c55e); }
.kpi-card.v1:hover { border-color: #93c5fd; }
.kpi-card.v2:hover { border-color: #67e8f9; }
.kpi-card.v3:hover { border-color: #fca5a5; }
.kpi-card.v4:hover { border-color: #86efac; }
.kpi-icon {
    font-size: 1.4rem;
    margin-bottom: 0.35rem;
}
.kpi-label {
    font-size: 0.78rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #64748b;
    margin-bottom: 0.2rem;
    font-weight: 500;
}
.kpi-value {
    font-size: 1.85rem;
    font-weight: 700;
    color: #0f172a;
    line-height: 1.15;
    letter-spacing: -0.01em;
}

/* --- セクションヘッダー --- */
.section-header {
    display: flex;
    align-items: center;
    gap: 0.65rem;
    margin: 2rem 0 0.9rem 0;
    padding-bottom: 0.55rem;
    border-bottom: 2px solid #e2e8f0;
}
.section-header .icon {
    font-size: 1.1rem;
    width: 2rem;
    height: 2rem;
    display: flex;
    align-items: center;
    justify-content: center;
    border-radius: 8px;
    flex-shrink: 0;
}
.section-header .icon.purple { background: #ede9fe; }
.section-header .icon.blue   { background: #dbeafe; }
.section-header .icon.red    { background: #fee2e2; }
.section-header .icon.green  { background: #dcfce7; }
.section-header .icon.amber  { background: #fef3c7; }
.section-header h3 {
    margin: 0 !important;
    font-size: 1.05rem;
    font-weight: 600;
    color: #1e293b !important;
    letter-spacing: 0.01em;
}

/* --- ダウンロードエリア --- */
.download-area {
    background: #ffffff;
    border: 1px dashed #94a3b8;
    border-radius: 14px;
    padding: 1.4rem;
    text-align: center;
    margin: 1rem 0;
}
.download-area p {
    color: #64748b !important;
    font-size: 0.85rem;
    margin-top: 0.5rem;
}

/* --- データフレーム --- */
.stDataFrame {
    border-radius: 10px;
    overflow: hidden;
    border: 1px solid #e2e8f0;
}

/* --- バッジ --- */
.badge {
    display: inline-block;
    padding: 0.2rem 0.65rem;
    border-radius: 9999px;
    font-size: 0.78rem;
    font-weight: 600;
    letter-spacing: 0.02em;
}
.badge-ok   { background: #dcfce7; color: #166534; }
.badge-warn { background: #fef3c7; color: #92400e; }
.badge-crit { background: #fee2e2; color: #991b1b; }

/* --- ウェルカム画面 --- */
.welcome-area {
    text-align: center;
    padding: 4.5rem 1rem;
}
.welcome-area .glow-icon {
    font-size: 3.5rem;
    display: inline-block;
    margin-bottom: 1rem;
}
.welcome-area p {
    color: #64748b !important;
    font-size: 0.95rem;
    line-height: 1.8;
}
.welcome-area strong {
    color: #2563eb !important;
}

/* --- フッター --- */
.app-footer {
    text-align: center;
    color: #94a3b8;
    font-size: 0.78rem;
    margin-top: 3rem;
    padding: 1.2rem 0;
    border-top: 1px solid #e2e8f0;
    letter-spacing: 0.02em;
}

/* --- メッセージ系の微調整 --- */
.stAlert { border-radius: 10px !important; }
</style>
"""


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------
def categorize_aging(days: int) -> str:
    for lo, hi, label in AGING_BINS:
        if lo <= days <= hi:
            return label
    return "365日超"


def parse_expiry(sub_inv: str) -> pd.Timestamp | None:
    """Sub Inventory から賞味期限を抽出する。"""
    if not isinstance(sub_inv, str):
        return None
    m = re.search(r"SS?_(\d{6})$", sub_inv)
    if not m:
        return None
    digits = m.group(1)
    try:
        yy, mm, dd = int(digits[:2]), int(digits[2:4]), int(digits[4:6])
        return pd.Timestamp(year=2000 + yy, month=mm, day=dd)
    except ValueError:
        return None


def expiry_status(earliest_expiry: pd.Timestamp | None, today: pd.Timestamp) -> str:
    if earliest_expiry is None or pd.isna(earliest_expiry):
        return ""
    if earliest_expiry <= today:
        return "期限切れ"
    if earliest_expiry <= today + timedelta(days=90):
        return "3ヶ月以内"
    return "期限あり"


def strip_leading_zeros(s: str) -> str:
    return s.lstrip("0")


# ---------------------------------------------------------------------------
# データ読み込み
# ---------------------------------------------------------------------------
def load_inventory(file) -> pd.DataFrame:
    try:
        df = pd.read_excel(file, engine="openpyxl")
    except Exception as e:
        raise ValueError(
            f"在庫リストの読み込みに失敗しました。\n"
            f"Excel形式（.xlsx）のファイルを指定してください。\n"
            f"詳細: {e}"
        )
    required = ["Product Code", "PICKING KEY7", "Arrival Date", "Sub Inventory"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"在庫リストに必要なカラムが見つかりません: {', '.join(missing)}\n"
            f"1行目がヘッダー行のExcelファイルか確認してください。"
        )
    return df


def load_shopee_files(files) -> pd.DataFrame:
    frames = []
    for f in files:
        try:
            df = pd.read_excel(f, skiprows=3, header=None, engine="calamine")
        except Exception as e:
            raise ValueError(
                f"Shopeeファイル「{f.name}」の読み込みに失敗しました。\n"
                f"Shopee管理画面からエクスポートしたExcelファイルか確認してください。\n"
                f"詳細: {e}"
            )
        if len(df.columns) >= len(SHOPEE_COLUMNS):
            df = df.iloc[:, : len(SHOPEE_COLUMNS)]
            df.columns = SHOPEE_COLUMNS
        else:
            df.columns = SHOPEE_COLUMNS[: len(df.columns)]
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.dropna(subset=["Product ID"])
    return combined


# ---------------------------------------------------------------------------
# Shopee 掲載判定
# ---------------------------------------------------------------------------
def build_shopee_sets(shopee_df: pd.DataFrame):
    sku_set = set(shopee_df["SKU"].dropna().astype(str).str.strip())
    gtin_set = set(shopee_df["GTIN"].dropna().astype(str).str.strip())
    barcode_set: set[str] = set()
    for sku in sku_set:
        parts = sku.split("_")
        if len(parts) >= 3:
            barcode = "_".join(parts[1:-1])
            barcode_set.add(barcode)
            barcode_set.add(strip_leading_zeros(barcode))
    return sku_set, gtin_set, barcode_set


def is_on_shopee(row: pd.Series, sku_set: set, gtin_set: set, barcode_set: set) -> bool:
    pk1 = str(row.get("PICKING KEY1", "")).strip()
    pcode = str(row.get("Product Code", "")).strip()
    if pk1 and pk1 in sku_set:
        return True
    if pcode and pcode in gtin_set:
        return True
    if pcode and (pcode in barcode_set or strip_leading_zeros(pcode) in barcode_set):
        return True
    return False


# ---------------------------------------------------------------------------
# メイン分析処理
# ---------------------------------------------------------------------------
def run_analysis(
    inv_df: pd.DataFrame,
    shopee_df: pd.DataFrame | None,
    include_blank_key7: bool = False,
) -> pd.DataFrame:
    today = pd.Timestamp(datetime.today().date())

    key7 = inv_df["PICKING KEY7"].astype(str).str.strip().str.upper()
    if include_blank_key7:
        mask = (key7 == "EC") | (key7.isin(["", "NAN", "NONE"]))
    else:
        mask = key7 == "EC"
    df = inv_df[mask].copy()
    if df.empty:
        st.error("対象レコードが見つかりません。PICKING KEY7 の値を確認してください。")
        return pd.DataFrame()

    df["賞味期限"] = df["Sub Inventory"].apply(parse_expiry)

    if shopee_df is not None and not shopee_df.empty:
        sku_set, gtin_set, barcode_set = build_shopee_sets(shopee_df)
        df["Shopee掲載"] = df.apply(lambda r: is_on_shopee(r, sku_set, gtin_set, barcode_set), axis=1)
    else:
        df["Shopee掲載"] = False

    df["Arrival Date"] = pd.to_datetime(df["Arrival Date"], errors="coerce")
    for col in ["Total Piece Qty", "Case Qty", "Total Weight", "Total Volume"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    grouped = df.groupby("Product Code", as_index=False).agg(
        商品名=("Product Name", "first"),
        入庫回数=("Arrival Date", "count"),
        最古入庫日=("Arrival Date", "min"),
        最新入庫日=("Arrival Date", "max"),
        合計数量=("Total Piece Qty", "sum"),
        合計ケース数=("Case Qty", "sum"),
        合計重量=("Total Weight", "sum"),
        合計体積=("Total Volume", "sum"),
        Shopee掲載=("Shopee掲載", "any"),
        最早期限日=("賞味期限", "min"),
        期限一覧=("賞味期限", lambda x: ", ".join(sorted(set(
            d.strftime("%Y-%m-%d") for d in x.dropna()
        )))),
    )

    grouped["滞留日数"] = (today - grouped["最古入庫日"]).dt.days
    grouped["滞留日数"] = grouped["滞留日数"].fillna(0).astype(int)
    grouped["Agingカテゴリ"] = grouped["滞留日数"].apply(categorize_aging)
    grouped["期限ステータス"] = grouped["最早期限日"].apply(lambda d: expiry_status(d, today))
    grouped["B2B候補"] = (grouped["滞留日数"] >= 90) | (grouped["合計数量"] >= 10)
    grouped = grouped.sort_values("滞留日数", ascending=False).reset_index(drop=True)
    return grouped


# ---------------------------------------------------------------------------
# Excel 出力
# ---------------------------------------------------------------------------
def _apply_header_style(ws, max_col: int):
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                val = str(cell.value) if cell.value is not None else ""
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _color_detail_rows(ws, header_map: dict, row_count: int):
    shopee_col = header_map.get("Shopee掲載")
    expiry_col = header_map.get("期限ステータス")
    aging_col = header_map.get("滞留日数")

    for row_idx in range(2, row_count + 2):
        if shopee_col:
            val = ws.cell(row=row_idx, column=shopee_col).value
            if val is True or str(val).strip() in ("True", "●", "1"):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=c).fill = FILL_SHOPEE

        if expiry_col:
            exp_val = str(ws.cell(row=row_idx, column=expiry_col).value or "")
            if exp_val == "期限切れ":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=c).fill = FILL_EXPIRED
                    ws.cell(row=row_idx, column=c).font = FONT_EXPIRED
            elif exp_val == "3ヶ月以内":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=c).fill = FILL_NEAR_EXPIRY

        if aging_col and expiry_col:
            exp_val = str(ws.cell(row=row_idx, column=expiry_col).value or "")
            if exp_val not in ("期限切れ", "3ヶ月以内"):
                days_val = ws.cell(row=row_idx, column=aging_col).value
                if isinstance(days_val, (int, float)):
                    days_int = int(days_val)
                    if days_int <= 60:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=c).fill = FILL_GREEN
                    elif days_int <= 180:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=c).fill = FILL_YELLOW
                    elif days_int > 180:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=c).fill = FILL_PINK


def _write_df_to_sheet(ws, df: pd.DataFrame, freeze: bool = True):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
                cell.number_format = "YYYY-MM-DD"
            elif isinstance(val, bool):
                cell.value = val
            else:
                cell.value = val
            cell.border = THIN_BORDER
    _apply_header_style(ws, len(df.columns))
    _auto_width(ws)
    if freeze:
        ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = ws.dimensions


def generate_excel(result_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    today_str = datetime.today().strftime("%Y-%m-%d")

    # --- シート1: サマリ ---
    ws1 = wb.active
    ws1.title = "サマリ"
    aging_summary = result_df.groupby("Agingカテゴリ", sort=False).agg(
        SKU数=("Product Code", "count"),
        Shopee掲載数=("Shopee掲載", "sum"),
        合計数量=("合計数量", "sum"),
        期限注意=("期限ステータス", lambda x: ((x == "期限切れ") | (x == "3ヶ月以内")).sum()),
    ).reset_index()
    aging_summary["構成比"] = (aging_summary["SKU数"] / aging_summary["SKU数"].sum() * 100).round(1)
    cat_order = [label for _, _, label in AGING_BINS]
    aging_summary["_sort"] = aging_summary["Agingカテゴリ"].apply(
        lambda x: cat_order.index(x) if x in cat_order else 999
    )
    aging_summary = aging_summary.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)

    ws1.append([f"在庫Aging分析サマリ（{today_str}）"])
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws1.cell(1, 1).font = Font(bold=True, size=14)
    ws1.append([])

    total_sku = len(result_df)
    shopee_count = int(result_df["Shopee掲載"].sum())
    expiry_warn = int(((result_df["期限ステータス"] == "期限切れ") | (result_df["期限ステータス"] == "3ヶ月以内")).sum())
    b2b_count = int(result_df["B2B候補"].sum())
    ws1.append(["全SKU数", total_sku, "", "Shopee掲載数", shopee_count])
    ws1.append(["期限注意数", expiry_warn, "", "B2B候補数", b2b_count])
    ws1.append([])

    ws1.append(["【Agingカテゴリ別集計】"])
    ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=11)
    start_row = ws1.max_row + 1
    headers = ["Agingカテゴリ", "SKU数", "Shopee掲載数", "合計数量", "期限注意", "構成比(%)"]
    ws1.append(headers)
    for c_idx in range(1, len(headers) + 1):
        cell = ws1.cell(start_row, c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for _, arow in aging_summary.iterrows():
        ws1.append([
            arow["Agingカテゴリ"], int(arow["SKU数"]), int(arow["Shopee掲載数"]),
            arow["合計数量"], int(arow["期限注意"]), arow["構成比"],
        ])

    ws1.append([])
    ws1.append(["【凡例】"])
    ws1.append(["水色行", "Shopee掲載済み"])
    ws1.append(["赤行", "期限切れ"])
    ws1.append(["オレンジ行", "期限3ヶ月以内"])
    ws1.append(["緑行", "Aging 0-60日"])
    ws1.append(["黄行", "Aging 61-180日"])
    ws1.append(["ピンク行", "Aging 181日超"])
    _auto_width(ws1)

    # --- シート2: 商品別Aging明細 ---
    ws2 = wb.create_sheet("商品別Aging明細")
    display_df = result_df.copy()
    display_df["Shopee掲載"] = display_df["Shopee掲載"].map({True: "●", False: ""})
    display_df["B2B候補"] = display_df["B2B候補"].map({True: "●", False: ""})
    _write_df_to_sheet(ws2, display_df)
    header_map = {col: i + 1 for i, col in enumerate(display_df.columns)}
    _color_detail_rows(ws2, {
        "Shopee掲載": header_map.get("Shopee掲載"),
        "期限ステータス": header_map.get("期限ステータス"),
        "滞留日数": header_map.get("滞留日数"),
    }, len(display_df))

    # --- シート3: 期限注意リスト ---
    ws3 = wb.create_sheet("⚠期限注意リスト")
    expiry_df = result_df[result_df["期限ステータス"].isin(["期限切れ", "3ヶ月以内"])].copy()
    expiry_df["Shopee掲載"] = expiry_df["Shopee掲載"].map({True: "●", False: ""})
    expiry_df["B2B候補"] = expiry_df["B2B候補"].map({True: "●", False: ""})
    if not expiry_df.empty:
        _write_df_to_sheet(ws3, expiry_df)
        hm3 = {col: i + 1 for i, col in enumerate(expiry_df.columns)}
        _color_detail_rows(ws3, {
            "Shopee掲載": hm3.get("Shopee掲載"),
            "期限ステータス": hm3.get("期限ステータス"),
            "滞留日数": hm3.get("滞留日数"),
        }, len(expiry_df))
    else:
        ws3.append(["期限注意の商品はありません。"])

    # --- シート4: B2B候補_Shopee未掲載 ---
    ws4 = wb.create_sheet("B2B候補_Shopee未掲載")
    b2b_df = result_df[(result_df["B2B候補"]) & (~result_df["Shopee掲載"])].copy()
    b2b_df["Shopee掲載"] = b2b_df["Shopee掲載"].map({True: "●", False: ""})
    b2b_df["B2B候補"] = b2b_df["B2B候補"].map({True: "●", False: ""})
    if not b2b_df.empty:
        _write_df_to_sheet(ws4, b2b_df)
        hm4 = {col: i + 1 for i, col in enumerate(b2b_df.columns)}
        _color_detail_rows(ws4, {
            "Shopee掲載": hm4.get("Shopee掲載"),
            "期限ステータス": hm4.get("期限ステータス"),
            "滞留日数": hm4.get("滞留日数"),
        }, len(b2b_df))
    else:
        ws4.append(["B2B候補（Shopee未掲載）の商品はありません。"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(result_df: pd.DataFrame) -> str:
    """スプレッドシート用 CSV を生成する。"""
    out = result_df.copy()
    out["Shopee掲載"] = out["Shopee掲載"].map({True: "●", False: ""})
    out["B2B候補"] = out["B2B候補"].map({True: "●", False: ""})
    for col in ["最古入庫日", "最新入庫日", "最早期限日"]:
        if col in out.columns:
            out[col] = out[col].apply(
                lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else ""
            )
    return out.to_csv(index=False)


# ---------------------------------------------------------------------------
# Slack 通知
# ---------------------------------------------------------------------------
def _build_summary_text(result_df: pd.DataFrame) -> str:
    """Slack 投稿用のサマリテキスト（ファイルと一緒に投稿するメッセージ）。"""
    today_str = datetime.today().strftime("%Y-%m-%d %H:%M")
    total_sku = len(result_df)
    shopee_count = int(result_df["Shopee掲載"].sum())
    expiry_warn = int(((result_df["期限ステータス"] == "期限切れ") | (result_df["期限ステータス"] == "3ヶ月以内")).sum())
    b2b_count = int(result_df["B2B候補"].sum())

    cat_order = [label for _, _, label in AGING_BINS]
    aging_lines = []
    for cat in cat_order:
        cnt = int((result_df["Agingカテゴリ"] == cat).sum())
        if cnt > 0:
            aging_lines.append(f"    {cat}: {cnt:,} SKU")
    aging_text = "\n".join(aging_lines) if aging_lines else "    データなし"

    return (
        f"📦 *在庫Aging分析レポート*\n"
        f"分析日時: {today_str}\n\n"
        f"*KPI サマリ*\n"
        f"    🏷 全SKU数: {total_sku:,}\n"
        f"    🛒 Shopee掲載: {shopee_count:,}\n"
        f"    ⚠️ 期限注意: {expiry_warn:,}\n"
        f"    📦 B2B候補: {b2b_count:,}\n\n"
        f"*📈 Aging 内訳*\n{aging_text}\n\n"
        f"_Excelファイルを添付しました。詳細はファイルをご確認ください。_"
    )


def send_slack_notification(
    bot_token: str, channel_id: str, result_df: pd.DataFrame, excel_bytes: bytes,
) -> tuple[bool, str]:
    """Slack Bot Token で Excel ファイル + サマリメッセージを送信する。"""
    token = bot_token.strip()
    ch = channel_id.strip()
    if not token.startswith("xoxb-"):
        return False, "Bot Token が正しくありません。xoxb- で始まるトークンを入力してください"
    if not ch:
        return False, "チャンネルIDが未入力です"

    today_str = datetime.today().strftime("%Y%m%d_%H%M")
    filename = f"aging_report_{today_str}.xlsx"
    summary = _build_summary_text(result_df)

    # --- Step 1: files.getUploadURLExternal で署名付きURLを取得 ---
    params = json.dumps({"filename": filename, "length": len(excel_bytes)}).encode()
    try:
        req = Request(
            f"https://slack.com/api/files.getUploadURLExternal"
            f"?filename={filename}&length={len(excel_bytes)}",
            method="GET",
            headers={"Authorization": f"Bearer {token}"},
        )
        with urlopen(req, timeout=15) as resp:
            body = json.loads(resp.read().decode())
        if not body.get("ok"):
            return False, f"Slack API エラー (getUploadURL): {body.get('error', body)}"
        upload_url = body["upload_url"]
        file_id = body["file_id"]
    except Exception as e:
        return False, f"アップロードURL取得に失敗: {e}"

    # --- Step 2: upload_url に PUT でファイルを送信 ---
    try:
        req2 = Request(
            upload_url,
            data=excel_bytes,
            method="POST",
            headers={"Content-Type": "application/octet-stream"},
        )
        with urlopen(req2, timeout=30) as resp2:
            if resp2.status not in (200, 201):
                return False, f"ファイルアップロード失敗: status={resp2.status}"
    except Exception as e:
        return False, f"ファイルアップロード失敗: {e}"

    # --- Step 3: files.completeUploadExternal でチャンネルに共有 ---
    try:
        complete_payload = json.dumps({
            "files": [{"id": file_id, "title": filename}],
            "channel_id": ch,
            "initial_comment": summary,
        }).encode()
        req3 = Request(
            "https://slack.com/api/files.completeUploadExternal",
            data=complete_payload,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json; charset=utf-8",
            },
        )
        with urlopen(req3, timeout=15) as resp3:
            body3 = json.loads(resp3.read().decode())
        if not body3.get("ok"):
            return False, f"Slack API エラー (completeUpload): {body3.get('error', body3)}"
    except Exception as e:
        return False, f"ファイル共有に失敗: {e}"

    return True, "Slack にExcelファイル + サマリを送信しました"


# ---------------------------------------------------------------------------
# UI ヘルパー
# ---------------------------------------------------------------------------
def render_section_header(icon: str, title: str, color: str = "blue"):
    st.markdown(
        f'<div class="section-header">'
        f'<span class="icon {color}">{icon}</span>'
        f'<h3>{title}</h3>'
        f'</div>',
        unsafe_allow_html=True,
    )


def render_kpi_cards(total_sku: int, shopee_count: int, expiry_warn: int, b2b_count: int):
    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card v1">
            <div class="kpi-icon">🏷</div>
            <div class="kpi-label">全SKU数</div>
            <div class="kpi-value">{total_sku:,}</div>
        </div>
        <div class="kpi-card v2">
            <div class="kpi-icon">🛒</div>
            <div class="kpi-label">Shopee掲載数</div>
            <div class="kpi-value">{shopee_count:,}</div>
        </div>
        <div class="kpi-card v3">
            <div class="kpi-icon">⚠️</div>
            <div class="kpi-label">期限注意数</div>
            <div class="kpi-value">{expiry_warn:,}</div>
        </div>
        <div class="kpi-card v4">
            <div class="kpi-icon">📦</div>
            <div class="kpi-label">B2B候補数</div>
            <div class="kpi-value">{b2b_count:,}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="在庫Aging分析", page_icon="📦", layout="wide")
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

    # ヘッダー
    st.markdown(
        "<h1 style='margin-bottom:0; font-weight:700; letter-spacing:-0.02em'>"
        "<span style='background:linear-gradient(135deg,#2563eb,#0891b2);-webkit-background-clip:text;-webkit-text-fill-color:transparent'>"
        "在庫 Aging 分析</span></h1>"
        "<p style='color:#64748b; margin-top:0.3rem; font-size:0.85rem; font-weight:400'>"
        "ユアトレード EC在庫 — Aging分析・期限管理・B2B候補抽出</p>",
        unsafe_allow_html=True,
    )

    cat_order = [label for _, _, label in AGING_BINS]

    # --- サイドバー ---
    with st.sidebar:
        st.markdown("### 📂 ファイル")
        inv_file = st.file_uploader(
            "在庫リスト（Excel）",
            type=["xlsx", "xls"],
            accept_multiple_files=False,
            key="inv",
        )
        shopee_files = st.file_uploader(
            "Shopee商品リスト（複数可）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="shopee",
        )

        st.markdown("---")
        st.markdown("### ⚙ オプション")
        include_blank_key7 = st.checkbox(
            "KEY7 空欄も含める",
            value=False,
            help="PICKING KEY7 が空欄の行も分析対象に含めます",
        )

        st.markdown("---")
        st.markdown("### 🔍 明細フィルタ")
        aging_filter = st.multiselect(
            "Aging カテゴリ", options=cat_order, default=cat_order, key="aging_filter",
        )
        shopee_filter = st.selectbox(
            "Shopee掲載", ["すべて", "掲載あり", "未掲載"], key="shopee_filter",
        )
        b2b_filter = st.selectbox(
            "B2B候補", ["すべて", "候補のみ", "候補外"], key="b2b_filter",
        )

        st.markdown("---")
        st.markdown("### 📤 Slack 共有")
        _env_bot_token = os.getenv("SLACK_BOT_TOKEN", "")
        _env_channel_id = os.getenv("SLACK_CHANNEL_ID", "")
        slack_bot_token = st.text_input(
            "Bot Token",
            value=_env_bot_token,
            type="password",
            placeholder="xoxb-...",
            help="Slack App の Bot User OAuth Token (xoxb-...)",
            key="slack_bot_token",
        )
        slack_channel_id = st.text_input(
            "チャンネルID",
            value=_env_channel_id,
            placeholder="C0XXXXXXX",
            help="送信先チャンネルのID（チャンネル名ではなくIDを入力）",
            key="slack_channel_id",
        )
        if _env_bot_token:
            st.caption("✅ .env から読み込み済み")

        st.markdown("---")
        run_btn = st.button("▶  分析実行", type="primary", use_container_width=True)

    # --- 分析結果を session_state に保持 ---
    if "result" not in st.session_state:
        st.session_state["result"] = None

    if run_btn:
        if inv_file is None:
            st.error("在庫リストをアップロードしてください。")
            return

        # データ読み込み
        try:
            with st.spinner("在庫リストを読み込み中..."):
                inv_df = load_inventory(inv_file)
        except ValueError as e:
            st.error(str(e))
            return
        except Exception as e:
            st.error(
                f"在庫リストの読み込み中に予期しないエラーが発生しました。\n\n"
                f"ファイルが正しい Excel 形式（.xlsx）か確認してください。\n\n詳細: {e}"
            )
            return

        shopee_df = None
        if shopee_files:
            try:
                with st.spinner("Shopee商品リストを読み込み中..."):
                    shopee_df = load_shopee_files(shopee_files)
                    st.sidebar.success(f"Shopee: {len(shopee_df):,}件")
            except ValueError as e:
                st.error(str(e))
                return
            except Exception as e:
                st.error(
                    f"Shopee商品リストの読み込み中にエラーが発生しました。\n\n"
                    f"Shopee管理画面からエクスポートした Excel ファイルか確認してください。\n\n詳細: {e}"
                )
                return

        with st.spinner("分析処理中..."):
            result = run_analysis(inv_df, shopee_df, include_blank_key7=include_blank_key7)

        if result.empty:
            st.warning("分析結果が0件です。入力データを確認してください。")
            return

        st.session_state["result"] = result

    # --- session_state から結果を取得して表示 ---
    result = st.session_state.get("result")
    if result is None:
        st.markdown(
            '<div class="welcome-area">'
            '<div class="glow-icon">📦</div>'
            '<p>サイドバーからファイルをアップロードし、<br>'
            '<strong>「分析実行」</strong>を押してください</p></div>',
            unsafe_allow_html=True,
        )
        return

    # =========================================
    # 1. KPI カード
    # =========================================
    total_sku = len(result)
    shopee_count = int(result["Shopee掲載"].sum())
    expiry_warn = int(((result["期限ステータス"] == "期限切れ") | (result["期限ステータス"] == "3ヶ月以内")).sum())
    b2b_count = int(result["B2B候補"].sum())
    render_kpi_cards(total_sku, shopee_count, expiry_warn, b2b_count)

    # =========================================
    # 2. Aging カテゴリ別集計
    # =========================================
    render_section_header("📈", "Aging カテゴリ別集計", "purple")
    aging_summary = result.groupby("Agingカテゴリ", sort=False).agg(
        SKU数=("Product Code", "count"),
        Shopee掲載数=("Shopee掲載", "sum"),
        合計数量=("合計数量", "sum"),
        期限注意=("期限ステータス", lambda x: ((x == "期限切れ") | (x == "3ヶ月以内")).sum()),
    ).reset_index()
    aging_summary["Shopee掲載数"] = aging_summary["Shopee掲載数"].astype(int)
    aging_summary["期限注意"] = aging_summary["期限注意"].astype(int)
    aging_summary["構成比(%)"] = (aging_summary["SKU数"] / aging_summary["SKU数"].sum() * 100).round(1)
    aging_summary["_sort"] = aging_summary["Agingカテゴリ"].apply(
        lambda x: cat_order.index(x) if x in cat_order else 999
    )
    aging_summary = aging_summary.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)
    st.dataframe(aging_summary, use_container_width=True, hide_index=True)

    # =========================================
    # 3. 期限注意リスト
    # =========================================
    render_section_header("🚨", "期限注意リスト", "red")
    expiry_df = result[result["期限ステータス"].isin(["期限切れ", "3ヶ月以内"])].copy()
    if expiry_df.empty:
        st.markdown(
            '<span class="badge badge-ok">OK — 期限注意の商品はありません</span>',
            unsafe_allow_html=True,
        )
    else:
        def highlight_expiry(row):
            if row["期限ステータス"] == "期限切れ":
                return ["background-color: #FF6B6B; color: white"] * len(row)
            if row["期限ステータス"] == "3ヶ月以内":
                return ["background-color: #FFA500"] * len(row)
            return [""] * len(row)

        display_exp = expiry_df[
            ["Product Code", "商品名", "合計数量", "最早期限日", "期限ステータス", "Shopee掲載", "滞留日数"]
        ].copy()
        display_exp["Shopee掲載"] = display_exp["Shopee掲載"].map({True: "●", False: ""})
        styled = display_exp.style.apply(highlight_expiry, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True)

    # =========================================
    # 4. 商品別 Aging 明細
    # =========================================
    render_section_header("📋", "商品別 Aging 明細", "blue")

    filtered = result.copy()
    filtered = filtered[filtered["Agingカテゴリ"].isin(aging_filter)]
    if shopee_filter == "掲載あり":
        filtered = filtered[filtered["Shopee掲載"]]
    elif shopee_filter == "未掲載":
        filtered = filtered[~filtered["Shopee掲載"]]
    if b2b_filter == "候補のみ":
        filtered = filtered[filtered["B2B候補"]]
    elif b2b_filter == "候補外":
        filtered = filtered[~filtered["B2B候補"]]

    display_full = filtered.copy()
    display_full["Shopee掲載"] = display_full["Shopee掲載"].map({True: "●", False: ""})
    display_full["B2B候補"] = display_full["B2B候補"].map({True: "●", False: ""})
    display_full["期限注意"] = display_full["期限ステータス"].apply(
        lambda x: "⚠" if x in ("期限切れ", "3ヶ月以内") else ""
    )
    st.dataframe(display_full, use_container_width=True, hide_index=True, height=500)
    st.caption(f"表示中: {len(filtered):,}件 / 全{len(result):,}件")

    # =========================================
    # 5. ダウンロード
    # =========================================
    render_section_header("💾", "ダウンロード", "green")
    today_str = datetime.today().strftime("%Y%m%d")

    st.markdown('<div class="download-area">', unsafe_allow_html=True)
    dl1, dl2 = st.columns(2)
    with dl1:
        excel_data = generate_excel(result)
        st.download_button(
            label="📥 Excel (.xlsx)",
            data=excel_data,
            file_name=f"在庫Aging分析_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with dl2:
        csv_data = generate_csv(result)
        st.download_button(
            label="📊 スプレッドシート用 CSV",
            data=csv_data.encode("utf-8-sig"),
            file_name=f"在庫Aging分析_{today_str}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    st.markdown(
        "<p>CSV は Google ドライブにアップロードし"
        "「Google スプレッドシートで開く」で利用できます</p>",
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    # =========================================
    # 6. Slack 共有
    # =========================================
    render_section_header("📤", "Slack に共有", "amber")
    if not slack_bot_token or not slack_channel_id:
        st.markdown(
            '<p style="color:#64748b; font-size:0.9rem">'
            'サイドバーの「Slack 共有」に Bot Token と チャンネルID を設定すると、'
            'ここからExcelファイル + サマリをチャンネルに送信できます</p>',
            unsafe_allow_html=True,
        )
    else:
        share_col1, share_col2 = st.columns([1, 2])
        with share_col1:
            share_btn = st.button(
                "🚀 Slack に送信",
                use_container_width=True,
                type="primary",
                key="slack_send",
            )
        with share_col2:
            st.markdown(
                '<p style="color:#64748b; font-size:0.85rem; margin-top:0.5rem">'
                'Excelファイル + KPIサマリ・Aging内訳がチャンネルに送信されます</p>',
                unsafe_allow_html=True,
            )
        if share_btn:
            with st.spinner("Slack にファイルを送信中..."):
                excel_data = generate_excel(result)
                ok, msg = send_slack_notification(
                    slack_bot_token, slack_channel_id, result, excel_data,
                )
            if ok:
                st.success(msg)
            else:
                st.error(msg)

    # フッター
    st.markdown(
        f'<div class="app-footer">'
        f'INVENTORY AGING ANALYZER v1.0 &nbsp;&middot;&nbsp; {datetime.today().strftime("%Y-%m-%d %H:%M")}'
        f'</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
